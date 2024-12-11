from openpyxl import Workbook
from openpyxl.styles.fonts import Font
from openpyxl.styles.alignment import Alignment

from automation import AutomationsSet
from configuration import ModuleSettingsLight
from const import DATA_FILES_ADDON_DIR, DATA_FILES_DIR
import psutil

header_font = Font(b=True, sz=14.0, color="c0372d")
subheader_font = Font(b=True, sz=12.0)
subheader_font_red = Font(b=True, sz=12.0, color="c0372d")
left_aligned = Alignment(horizontal="left", vertical="top", wrapText=True)


def create_documentation(router, filename):
    """Take settings information and create full domumentation."""
    page = add_css_info()
    doc = Workbook()
    page = document_overview(doc, page, router, router.modules)
    page = document_hub(doc, page, router.api_srv.sm_hub)
    page = document_router(doc, page, router)

    for idx in range(len(router.modules)):
        mod = router.modules[idx]
        # ws.add_image("./web/configurator_files/logo.png", "F1")
        page = document_module(doc, page, mod, idx)
    page += "</body>\n"

    del doc["Sheet"]
    if router.api_srv.is_addon:
        data_file_path = DATA_FILES_ADDON_DIR
    else:
        data_file_path = DATA_FILES_DIR
    doc.save(data_file_path + filename)
    return page


def document_overview(doc, page, rt, mods) -> str:
    """Create system overview in excel sheet."""

    channel_str = {1: "1 + 2", 2: "3 + 4", 3: "5 + 6", 4: "7 + 8"}
    hub = rt.api_srv.sm_hub

    page += '      <h1 id="overview">Systemübersicht</h1>\n'
    page += "      <h2>Zentrale</h2>\n"
    page += "      <table>\n"
    page += "        <thead>\n"
    page += "          <tr>\n"
    page += "            <th>Typ</th>\n"
    page += "            <th>Name</th>\n"
    page += "          </tr>\n"
    page += "        </thead>\n"
    page += "        <tbody>\n"
    page += "          <tr>\n"
    page += "            <td>Hub</td>\n"
    page += f'            <td><a href="#hub">{hub._host}</a></td>\n'
    page += "          </tr>\n"
    page += "          <tr>\n"
    page += "            <td>Router</td>\n"
    page += f'            <td><a href="#rt">{clean_name(rt._name)}</a></td>\n'
    page += "          </tr>\n"
    page += "        </tbody>\n"
    page += "      </table>\n"
    page += "      <h2>Module</h2>\n"
    page += "      <table>\n"
    page += "        <thead>\n"
    page += "          <tr>\n"
    page += "            <th>Nr.</th>\n"
    page += "            <th>Name</th>\n"
    page += "            <th>Typ</th>\n"
    page += "            <th>Bereich</th>\n"
    page += "            <th>Adr.</th>\n"
    page += "            <th>Kanalpaar</th>\n"
    page += "          </tr>\n"
    page += "        </thead>\n"
    page += "        <tbody>\n"
    mod_cnt = 1
    for mod in mods:
        page += "          <tr>\n"
        page += f"            <td>{mod_cnt}</td>\n"
        page += f'            <td><a href="#mod_{mod._id}">{clean_name(mod._name)}</a></td>\n'
        page += f"            <td>{mod._type}</td>\n"
        page += f"            <td>{mod.get_area_name()}</td>\n"
        page += f"            <td>{mod._id}</td>\n"
        page += f"            <td>{channel_str[mod._channel]}</td>\n"
        page += "          </tr>\n"
        mod_cnt += 1
    page += "        </tbody>\n"
    page += "      </table>\n"

    ws = doc.create_sheet("System", 0)
    row = 1
    ws.cell(row, 1).value = "Systemübersicht"
    ws.cell(row, 1).font = header_font
    row += 2
    ws.cell(row, 1).value = "Nr."
    ws.cell(row, 2).value = "Name"
    ws.cell(row, 3).value = "Typ"
    ws.cell(row, 4).value = "Bereich"
    ws.cell(row, 5).value = "Adr."
    ws.cell(row, 6).value = "Kanal"
    ws.cell(row, 1).font = subheader_font
    ws.cell(row, 2).font = subheader_font
    ws.cell(row, 3).font = subheader_font
    ws.cell(row, 4).font = subheader_font
    ws.cell(row, 5).font = subheader_font
    ws.cell(row, 6).font = subheader_font
    row += 1
    mod_cnt = 1
    for mod in mods:
        ws.cell(row, 1).value = f"{mod_cnt}"
        ws.cell(row, 1).alignment = left_aligned
        ws.cell(row, 2).value = clean_name(mod._name)
        ws.cell(row, 3).value = mod._type
        ws.cell(row, 4).value = mod.get_area_name()
        ws.cell(row, 5).value = f"{mod._id}"
        ws.cell(row, 6).value = f"{mod._channel}"
        mod_cnt += 1
        row += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 28
    ws.set_printer_settings(ws.PAPERSIZE_A4, ws.ORIENTATION_LANDSCAPE)

    page += '    <br><br><a href="#overview">zur Übersicht</a><hr><div class="pagebreak"> </div>\n'
    return page


def document_hub(doc, page, hub) -> str:
    """Export router information to excel sheet."""
    ws = doc.create_sheet(hub._host, 1)

    if hub.lan_mac == hub.curr_mac:
        netw_type = "LAN"
    else:
        netw_type = "WLAN"

    memory = psutil.virtual_memory()
    mem_avail = round(memory.available / 1024.0 / 1024.0, 1)
    mem_total = round(memory.total / 1024.0 / 1024.0, 1)
    mem_desc = (
        f"{mem_total} MB, genutzt {round(100 - (100 * mem_avail / mem_total), 1)}%"
    )
    disk = psutil.disk_usage("/")
    dsk_free = round(disk.free / 1024.0 / 1024.0 / 1024.0, 1)
    dsk_total = round(disk.total / 1024.0 / 1024.0, 1)
    dsk_desc = (
        f"{dsk_total} GB, genutzt {round(100 - (100 * dsk_free / dsk_total), 1)}%"
    )

    page += '      <h1 id="hub">' + f"Hub '{hub._host}'</h1>\n"
    page += "      <table>\n"
    page += "        <thead>\n"
    page += "          <tr>\n"
    page += "            <th>Eigenschaften</th>\n"
    page += "          </tr>\n"
    page += "        </thead>\n"
    page += "      </table>\n"
    page += "      <table>\n"
    page += "        <tr>\n"
    page += "          <td>Firmware:</td>\n"
    page += f"          <td>{hub.get_version()}</td>\n"
    page += "        </tr>\n"
    page += "        <tr>\n"
    page += "          <td>Seriennnr.:</td>\n"
    page += f"          <td>{hub._serial}</td>\n"
    page += "        </tr>\n"
    page += "        <tr>\n"
    page += "          <td>Hardware:</td>\n"
    page += f"          <td>{hub._pi_model}</td>\n"
    page += "        </tr>\n"
    page += "        <tr>\n"
    page += "          <td>Hardware:</td>\n"
    page += f"          <td>{hub._pi_model}</td>\n"
    page += "        </tr>\n"
    page += "        <tr>\n"
    page += "          <td>Arbeitsspeicher:</td>\n"
    page += f"          <td>{mem_desc}</td>\n"
    page += "        </tr>\n"
    page += "        <tr>\n"
    page += "          <td>Dateispeicher:</td>\n"
    page += f"          <td>{dsk_desc}</td>\n"
    page += "        </tr>\n"
    page += "        <tr>\n"
    page += "          <td>IP-Adresse:</td>\n"
    page += f"          <td>{hub._host_ip}</td>\n"
    page += "        </tr>\n"
    page += "      </table>\n"
    row = 1

    ws.cell(row, 1).value = f"Router '{hub._host}'"
    ws.cell(row, 1).font = header_font
    row += 2
    ws.cell(row, 1).value = "Eigenschaften"
    ws.cell(row, 1).font = subheader_font_red
    row += 1
    ws.cell(row, 1).value = "Firmware"
    ws.cell(row, 2).value = hub.get_version()
    row += 1
    ws.cell(row, 1).value = "Seriennnr.:"
    ws.cell(row, 2).value = hub._serial
    row += 1
    ws.cell(row, 1).value = "Hardware"
    ws.cell(row, 2).value = hub._pi_model
    row += 1
    ws.cell(row, 1).value = "Arbeitsspeicher"
    ws.cell(row, 2).value = mem_desc
    row += 1
    ws.cell(row, 1).value = "Dateispeicher"
    ws.cell(row, 2).value = dsk_desc
    row += 1
    ws.cell(row, 1).value = "Netzwerk"
    ws.cell(row, 2).value = netw_type
    row += 1
    ws.cell(row, 1).value = "IP-Adresse"
    ws.cell(row, 2).value = hub._host_ip
    row += 1

    page += '    <br><br><a href="#overview">zur Übersicht</a><hr><div class="pagebreak"> </div>\n'

    return page


def document_router(doc, page, rt) -> str:
    """Export router information to excel sheet."""
    ws = doc.create_sheet(clean_name(rt._name), 2)
    rt_serial = rt.serial[1:].decode("iso8859-1").strip()
    settings = rt.settings

    page += '      <h1 id="rt">' + f"Router '{clean_name(rt._name)}'</h1>\n"
    page += "      <table>\n"
    page += "        <thead>\n"
    page += "          <tr>\n"
    page += "            <th>Eigenschaften</th>\n"
    page += "          </tr>\n"
    page += "        </thead>\n"
    page += "      </table>\n"
    page += "      <table>\n"
    page += "        <tr>\n"
    page += "          <td>Firmware:</td>\n"
    page += f"          <td>{rt.get_version()}</td>\n"
    page += "        </tr>\n"
    page += "        <tr>\n"
    page += "          <td>Seriennnr.:</td>\n"
    page += f"          <td>{rt_serial}</td>\n"
    page += "        </tr>\n"
    page += "        <tr>\n"
    page += "          <td>Benutzer Modus 1:</td>\n"
    page += f"          <td>{settings.user1_name}</td>\n"
    page += "        </tr>\n"
    page += "        <tr>\n"
    page += "          <td>Benutzer Modus 2:</td>\n"
    page += f"          <td>{settings.user2_name}</td>\n"
    page += "        </tr>\n"
    page += "      </table>\n"
    row = 1

    ws.cell(row, 1).value = f"Router '{rt._name}'"
    ws.cell(row, 1).font = header_font
    row += 2
    ws.cell(row, 1).value = "Eigenschaften"
    ws.cell(row, 1).font = subheader_font_red
    row += 1
    ws.cell(row, 1).value = "Firmware"
    ws.cell(row, 2).value = rt.get_version()
    row += 1
    ws.cell(row, 1).value = "Seriennnr.:"
    ws.cell(row, 2).value = rt_serial
    row += 1
    ws.cell(row, 1).value = "Benutzer Modus 1"
    ws.cell(row, 2).value = settings.user1_name
    row += 1
    ws.cell(row, 1).value = "Benutzer Modus 2"
    ws.cell(row, 2).value = settings.user2_name
    row += 1

    if len(settings.areas):
        page += "      <h2>Bereiche</h2>\n"
        page += "      <table>\n"
        page += "        <thead>\n"
        page += "          <tr>\n"
        page += "            <th>Nr.</th>\n"
        page += "            <th>Name</th>\n"
        page += "          </tr>\n"
        page += "        </thead>\n"
        page += "        <tbody>\n"

        ws.cell(row, 1).value = "Bereiche"
        ws.cell(row, 1).font = subheader_font_red
        row += 1
        ws = write_col_headers(ws, row)
        row += 1
        for area in settings.areas:
            page += "          <tr>\n"
            page += f"            <td>{area.nmbr}</td>\n"
            page += f"            <td>{area.name}</td>\n"
            page += "          </tr>\n"
            ws.cell(row, 1).value = area.nmbr
            ws.cell(row, 1).alignment = left_aligned
            ws.cell(row, 2).value = area.name
            row += 1
        page += "        </tbody>\n"
        page += "      </table>\n"
        row += 1

    if len(settings.groups):
        grp_type = {
            0: "unabhängig",
            1: "Tag/Nacht",
            2: "Alarm",
            3: "Tag/Nacht, Alarm",
        }
        page += "      <h2>Gruppen</h2>\n"
        page += "      <table>\n"
        page += "        <thead>\n"
        page += "          <tr>\n"
        page += "            <th>Nr.</th>\n"
        page += "            <th>Name</th>\n"
        page += "            <th>Modi von Gruppe 0</th>\n"
        page += "          </tr>\n"
        page += "        </thead>\n"
        page += "        <tbody>\n"

        ws.cell(row, 1).value = "Gruppen"
        ws.cell(row, 1).font = subheader_font_red
        row += 1
        ws = write_col_headers(ws, row)
        row += 1
        for grp in settings.groups:
            if grp.nmbr > 0:
                mod_dep = grp_type[settings.mode_dependencies[grp.nmbr - 1]]
            else:
                mod_dep = ""
            page += "          <tr>\n"
            page += f"            <td>{grp.nmbr}</td>\n"
            page += f"            <td>{grp.name}</td>\n"
            page += f"            <td>{mod_dep}</td>\n"
            page += "          </tr>\n"
            ws.cell(row, 1).value = grp.nmbr
            ws.cell(row, 1).alignment = left_aligned
            ws.cell(row, 2).value = grp.name
            ws.cell(row, 3).value = mod_dep
            row += 1
        page += "        </tbody>\n"
        page += "      </table>\n"
        row += 1

    if len(settings.glob_flags):
        page += "      <h2>Globale Merker</h2>\n"
        page += "      <table>\n"
        page += "        <thead>\n"
        page += "          <tr>\n"
        page += "            <th>Nr.</th>\n"
        page += "            <th>Name</th>\n"
        page += "          </tr>\n"
        page += "        </thead>\n"
        page += "        <tbody>\n"

        ws.cell(row, 1).value = "Globale Merker"
        ws.cell(row, 1).font = subheader_font_red
        row += 1
        ws = write_col_headers(ws, row)
        row += 1
        for flg in settings.glob_flags:
            page += "          <tr>\n"
            page += f"            <td>{flg.nmbr}</td>\n"
            page += f"            <td>{flg.name}</td>\n"
            page += "          </tr>\n"
            ws.cell(row, 1).value = flg.nmbr
            ws.cell(row, 1).alignment = left_aligned
            ws.cell(row, 2).value = flg.name
            row += 1
        page += "        </tbody>\n"
        page += "      </table>\n"
        row += 1

    if len(settings.coll_cmds):
        page += "      <h2>Sammelbefehle</h2>\n"
        page += "      <table>\n"
        page += "        <thead>\n"
        page += "          <tr>\n"
        page += "            <th>Nr.</th>\n"
        page += "            <th>Name</th>\n"
        page += "          </tr>\n"
        page += "        </thead>\n"
        page += "        <tbody>\n"

        ws.cell(row, 1).value = "Sammelbefehle"
        ws.cell(row, 1).font = subheader_font_red
        row += 1
        ws = write_col_headers(ws, row)
        row += 1
        for cmd in settings.coll_cmds:
            page += "          <tr>\n"
            page += f"            <td>{cmd.nmbr}</td>\n"
            page += f"            <td>{cmd.name}</td>\n"
            page += "          </tr>\n"
            ws.cell(row, 1).value = cmd.nmbr
            ws.cell(row, 1).alignment = left_aligned
            ws.cell(row, 2).value = cmd.name
            row += 1
        page += "        </tbody>\n"
        page += "      </table>\n"
        row += 1

    page += '    <br><br><a href="#overview">zur Übersicht</a><hr><div class="pagebreak"> </div>\n'

    return page


def document_module(doc, page, mod, idx) -> str:
    """Export module information to excel sheet."""

    ws = doc.create_sheet(clean_name(mod._name), idx + 3)
    input_type = {1: "Taster", 2: "Schalter", 3: "Analog"}
    output_type = {-10: "", 1: "", 2: "Dimmbar"}
    cover_type = {-1: "Rollladen", 1: "Rollladen", -2: "Jalousie", 2: "Jalousie"}
    channel_str = {1: "1 + 2", 2: "3 + 4", 3: "5 + 6", 4: "7 + 8"}
    settings = mod.settings
    automation_set = AutomationsSet(settings)

    page += f'      <h1 id="mod_{mod._id}">' + f"Modul '{clean_name(mod._name)}'</h1>\n"
    page += "      <table>\n"
    page += "        <thead>\n"
    page += "          <tr>\n"
    page += "            <th>Eigenschaften</th>\n"
    page += "          </tr>\n"
    page += "        </thead>\n"
    page += "      </table>\n"
    page += "      <table>\n"
    page += "        <tr>\n"
    page += "          <td>Typ:</td>\n"
    page += f"          <td>{mod._type}</td>\n"
    page += "          <td></td>\n"
    page += "          <td></td>\n"
    page += "        </tr>\n"
    page += "        <tr>\n"
    page += "          <td>Adresse:</td>\n"
    page += f"          <td>{mod._id}</td>\n"
    page += "          <td>Firmware:</td>\n"
    page += f"          <td>{mod.get_sw_version()}</td>\n"
    page += "        </tr>\n"
    page += "        <tr>\n"
    page += "          <td>Kanalpaar:</td>\n"
    page += f"          <td>{channel_str[mod._channel]}</td>\n"
    page += "          <td>Seriennnr.:</td>\n"
    page += f"          <td>{mod._serial}</td>\n"
    page += "        </tr>\n"
    page += "        <tr>\n"
    page += "          <td>Bereich:</td>\n"
    page += f"          <td>{mod.get_area_name()}</td>\n"
    page += "          <td></td>\n"
    page += "          <td></td>\n"
    page += "        </tr>\n"
    page += "      </table>\n"

    row = 1
    ws.cell(row, 1).value = f"Modul '{mod._name}'"
    ws.cell(row, 1).font = header_font
    row += 2
    ws.cell(row, 1).value = "Grundeinstellungen"
    ws.cell(row, 1).font = subheader_font_red
    row += 1
    ws.cell(row, 1).value = "Typ:"
    ws.cell(row, 2).value = mod._type
    row += 1
    ws.cell(row, 1).value = "Adresse:"
    ws.cell(row, 2).value = mod._id
    ws.cell(row, 2).alignment = left_aligned
    ws.cell(row, 3).value = "Firmware:"
    ws.cell(row, 4).value = mod.get_sw_version()
    row += 1
    ws.cell(row, 1).value = "Kanalp.:"
    ws.cell(row, 2).value = channel_str[mod._channel]
    ws.cell(row, 2).alignment = left_aligned
    ws.cell(row, 3).value = "Seriennnr.:"
    ws.cell(row, 4).value = mod._serial
    row += 1
    ws.cell(row, 1).value = "Bereich:"
    ws.cell(row, 2).value = mod.get_area_name()

    if settings is None:
        settings = ModuleSettingsLight(mod)
    row += 2
    if len(settings.inputs):
        page += "      <h2>Eingänge</h2>\n"
        page += "      <table>\n"
        page += "        <thead>\n"
        page += "          <tr>\n"
        page += "            <th>Nr.</th>\n"
        page += "            <th>Name</th>\n"
        page += "            <th>Typ</th>\n"
        page += "            <th>Konfiguration</th>\n"
        page += "            <th>Bereich</th>\n"
        page += "          </tr>\n"
        page += "        </thead>\n"
        page += "        <tbody>\n"

        ws.cell(row, 1).value = "Eingänge"
        ws.cell(row, 1).font = subheader_font_red
        row += 1
        ws = write_col_headers(ws, row)
        row += 1
        for inpt in settings.inputs:
            if mod._typ in [b"\x0a\x1e", b"\x0b\x1e", b"\x32\x01"]:
                type_desc = "24V"
            elif (
                mod._typ in [b"\x01\x01", b"\x01\x02", b"\x01\x03", b"\x0b\x1f"]
                and inpt.type == 3
            ):
                type_desc = "0..10V"
            elif mod._typ in [b"\x0b\x01"]:
                type_desc = "230V"
            elif inpt.nmbr in range(5, 11):
                type_desc = "24V"
            else:
                type_desc = "230V"
            page += "          <tr>\n"
            page += f"            <td>{inpt.nmbr}</td>\n"
            page += f"            <td>{clean_name(inpt.name)}</td>\n"
            page += f"            <td>{type_desc}</td>\n"
            page += f"            <td>{input_type[inpt.type]}</td>\n"
            page += f"            <td>{get_area_name(inpt, mod)}</td>\n"
            page += "          </tr>\n"
            ws.cell(row, 1).value = inpt.nmbr
            ws.cell(row, 1).alignment = left_aligned
            ws.cell(row, 2).value = clean_name(inpt.name)
            ws.cell(row, 3).value = type_desc
            ws.cell(row, 4).value = input_type[inpt.type]
            ws.cell(row, 5).value = get_area_name(inpt, mod)
            row += 1
        page += "        </tbody>\n"
        page += "      </table>\n"
        row += 1
    if len(settings.outputs):
        page += "      <h2>Ausgänge</h2>\n"
        page += "      <table>\n"
        page += "        <thead>\n"
        page += "          <tr>\n"
        page += "            <th>Nr.</th>\n"
        page += "            <th>Name</th>\n"
        page += "            <th>Typ</th>\n"
        page += "            <th>Konfiguration</th>\n"
        page += "            <th>Bereich</th>\n"
        page += "          </tr>\n"
        page += "        </thead>\n"
        page += "        <tbody>\n"
        ws.cell(row, 1).value = "Ausgänge"
        ws.cell(row, 1).font = subheader_font_red
        row += 1
        ws = write_col_headers(ws, row)
        row += 1
        for outpt in settings.outputs:
            if mod._typ in [b"\x0a\x01", b"\x0a\x1e", b"\x0a\x32", b"\x0a\x33"]:
                type_desc = "Relais"
            elif mod._typ in [b"\x32\x01"]:
                type_desc = "24V"
            elif outpt.nmbr == 15:
                type_desc = "Relais"
            elif outpt.nmbr in [13, 14]:
                type_desc = "24V"
            else:
                type_desc = "230V"
            page += "          <tr>\n"
            page += f"            <td>{outpt.nmbr}</td>\n"
            page += f"            <td>{clean_name(outpt.name)}</td>\n"
            page += f"            <td>{type_desc}</td>\n"
            page += f"            <td>{output_type[outpt.type]}</td>\n"
            page += f"            <td>{get_area_name(outpt, mod)}</td>\n"
            page += "          </tr>\n"
            ws.cell(row, 1).value = outpt.nmbr
            ws.cell(row, 1).alignment = left_aligned
            ws.cell(row, 2).value = clean_name(outpt.name)
            ws.cell(row, 3).value = type_desc
            ws.cell(row, 4).value = output_type[outpt.type]
            ws.cell(row, 5).value = get_area_name(outpt, mod)
            row += 1
        page += "        </tbody>\n"
        page += "      </table>\n"
        row += 1
    if len(settings.covers):
        page += "      <h2>Rollladen</h2>\n"
        page += "      <table>\n"
        page += "        <thead>\n"
        page += "          <tr>\n"
        page += "            <th>Nr.</th>\n"
        page += "            <th>Name</th>\n"
        page += "            <th>Typ</th>\n"
        page += "            <th>Konfiguration</th>\n"
        page += "            <th>Bereich</th>\n"
        page += "          </tr>\n"
        page += "        </thead>\n"
        page += "        <tbody>\n"
        ws.cell(row, 1).value = "Rollladen"
        ws.cell(row, 1).font = subheader_font_red
        row += 1
        ws = write_col_headers(ws, row)
        row += 1
        for cov in settings.covers:
            if cov.type != 0:
                if cov.type > 0:
                    polarity_desc = "Ausgang A: auf, B: zu"
                else:
                    polarity_desc = "Ausgang A: zu, B: auf"
                page += "          <tr>\n"
                page += f"            <td>{cov.nmbr}</td>\n"
                page += f"            <td>{clean_name(cov.name)}</td>\n"
                page += f"            <td>{cover_type[cov.type]}</td>\n"
                page += f"            <td>{polarity_desc}</td>\n"
                page += f"            <td>{get_area_name(cov, mod)}</td>\n"
                page += "          </tr>\n"
                ws.cell(row, 1).value = cov.nmbr
                ws.cell(row, 1).alignment = left_aligned
                ws.cell(row, 2).value = clean_name(cov.name)
                ws.cell(row, 3).value = cover_type[cov.type]
                ws.cell(row, 4).value = polarity_desc
                ws.cell(row, 5).value = get_area_name(cov, mod)
                row += 1
        page += "        </tbody>\n"
        page += "      </table>\n"
        row += 1

    if len(automation_set.local):
        page += "      <h2>Lokale Automatisierungen</h2>\n"
        page += "      <table>\n"
        page += "        <thead>\n"
        page += "          <tr>\n"
        page += "            <th>Nr.</th>\n"
        page += "            <th>Auslöser</th>\n"
        page += "            <th>Bedingung</th>\n"
        page += "            <th>Aktion</th>\n"
        page += "          </tr>\n"
        page += "        </thead>\n"
        page += "        <tbody>\n"
        ws.cell(row, 1).value = "Lokale Automatisierungen"
        ws.cell(row, 1).font = subheader_font_red
        row += 1
        ws = write_atm_headers(ws, row)
        row += 1
        atm_no = 1
        for atm in automation_set.local:
            page += "          <tr>\n"
            page += f"            <td>{atm_no}</td>\n"
            page += f"            <td>{atm.trigger.description}</td>\n"
            page += f"            <td>{atm.condition.name}</td>\n"
            page += f"            <td>{atm.action.description}</td>\n"
            page += "          </tr>\n"
            ws.cell(row, 1).value = atm_no
            ws.cell(row, 1).alignment = left_aligned
            ws.cell(row, 2).value = atm.trigger.description
            ws.cell(row, 2).alignment = left_aligned
            ws.cell(row, 3).value = atm.condition.name
            ws.cell(row, 3).alignment = left_aligned
            ws.cell(row, 4).value = atm.action.description
            ws.cell(row, 4).alignment = left_aligned
            atm_no += 1
            row += 1
        page += "        </tbody>\n"
        page += "      </table>\n"
        row += 1
    if len(automation_set.external):
        ext_mod_name = ""
        page += "      <h2>Externe Automatisierungen</h2>\n"
        page += "      <table>\n"
        page += "        <thead>\n"
        page += "          <tr>\n"
        page += "            <th>Nr.</th>\n"
        page += "            <th>Auslöser</th>\n"
        page += "            <th>Bedingung</th>\n"
        page += "            <th>Aktion</th>\n"
        page += "          </tr>\n"
        page += "        </thead>\n"
        page += "        <tbody>\n"
        ws.cell(row, 1).value = "Externe Automatisierungen"
        ws.cell(row, 1).font = subheader_font_red
        row += 1
        ws = write_atm_ext_headers(ws, row)
        row += 1
        for atm in automation_set.external:
            if mod.get_rtr().get_module(atm.src_mod) is None:
                curr_mod_name = f"Mod_{atm.src_mod}?"
            else:
                curr_mod_name = mod.get_rtr().get_module(atm.src_mod)._name
            if ext_mod_name != curr_mod_name:
                ext_mod_name = curr_mod_name
                atm_no = 1
                ws.cell(row, 2).value = f"Modul {atm.src_mod} '{ext_mod_name}'"
                ws.cell(row, 2).font = subheader_font
                row += 1
                page += "          <tr>\n"
                page += "            <th></th>\n"
                page += f"            <th>Modul {atm.src_mod} '{ext_mod_name}'</th>\n"
                page += "            <th></th>\n"
                page += "            <th></th>\n"
                page += "          </tr>\n"
            page += "          <tr>\n"
            page += f"            <td>{atm_no}</td>\n"
            page += f"            <td>{atm.trigger.description}</td>\n"
            page += f"            <td>{atm.condition.name}</td>\n"
            page += f"            <td>{atm.action.description}</td>\n"
            page += "          </tr>\n"
            ws.cell(row, 1).value = atm_no
            ws.cell(row, 1).alignment = left_aligned
            ws.cell(row, 2).value = atm.trigger.description
            ws.cell(row, 2).alignment = left_aligned
            ws.cell(row, 3).value = atm.condition.name
            ws.cell(row, 3).alignment = left_aligned
            ws.cell(row, 4).value = atm.action.description
            ws.cell(row, 4).alignment = left_aligned
            atm_no += 1
            row += 1
        page += "        </tbody>\n"
        page += "      </table>\n"

    if len(automation_set.forward):
        ext_mod_name = ""
        page += "      <h2>Weitergeleitete Automatisierungen</h2>\n"
        page += "      <table>\n"
        page += "        <thead>\n"
        page += "          <tr>\n"
        page += "            <th>Nr.</th>\n"
        page += "            <th>Auslöser</th>\n"
        page += "            <th>Bedingung</th>\n"
        page += "            <th>Aktion</th>\n"
        page += "          </tr>\n"
        page += "        </thead>\n"
        page += "        <tbody>\n"
        ws.cell(row, 1).value = "Weitergeleitete Automatisierungen"
        ws.cell(row, 1).font = subheader_font_red
        row += 1
        ws = write_atm_ext_headers(ws, row)
        row += 1
        for atm in automation_set.forward:
            if mod.get_rtr().get_module(atm.src_mod) is None:
                curr_mod_name = f"Mod_{atm.src_mod}?"
            else:
                curr_mod_name = mod.get_rtr().get_module(atm.src_mod)._name
            if ext_mod_name != curr_mod_name:
                ext_mod_name = curr_mod_name
                atm_no = 1
                ws.cell(row, 2).value = f"Modul {atm.src_mod} '{ext_mod_name}'"
                ws.cell(row, 2).font = subheader_font
                row += 1
                page += "          <tr>\n"
                page += "            <th></th>\n"
                page += f"            <th>Modul {atm.src_mod} '{ext_mod_name}'</th>\n"
                page += "            <th></th>\n"
                page += "            <th></th>\n"
                page += "          </tr>\n"
            page += "          <tr>\n"
            page += f"            <td>{atm_no}</td>\n"
            page += f"            <td>{atm.trigger.description}</td>\n"
            page += f"            <td>{atm.condition.name}</td>\n"
            page += f"            <td>{atm.action.description}</td>\n"
            page += "          </tr>\n"
            ws.cell(row, 1).value = atm_no
            ws.cell(row, 1).alignment = left_aligned
            ws.cell(row, 2).value = atm.trigger.description
            ws.cell(row, 2).alignment = left_aligned
            ws.cell(row, 3).value = atm.condition.name
            ws.cell(row, 3).alignment = left_aligned
            ws.cell(row, 4).value = atm.action.description
            ws.cell(row, 4).alignment = left_aligned
            atm_no += 1
            row += 1
        page += "        </tbody>\n"
        page += "      </table>\n"

    page += '    <br><br><a href="#overview">zur Übersicht</a><hr><div class="pagebreak"> </div>\n'

    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20

    return page


def write_col_headers(ws, row: int):
    """Set column headers."""

    ws.cell(row, 1).value = "Nr."
    ws.cell(row, 2).value = "Name"
    ws.cell(row, 3).value = "Typ"
    ws.cell(row, 4).value = "Konfiguration"
    ws.cell(row, 5).value = "Bereich"
    ws.cell(row, 1).font = subheader_font
    ws.cell(row, 2).font = subheader_font
    ws.cell(row, 3).font = subheader_font
    ws.cell(row, 4).font = subheader_font
    ws.cell(row, 5).font = subheader_font
    return ws


def write_atm_headers(ws, row: int):
    """Set column headers."""

    ws.cell(row, 1).value = "Nr."
    ws.cell(row, 2).value = "Auslöser"
    ws.cell(row, 3).value = "Bedingung"
    ws.cell(row, 4).value = "Aktion"
    ws.cell(row, 1).font = subheader_font
    ws.cell(row, 2).font = subheader_font
    ws.cell(row, 3).font = subheader_font
    ws.cell(row, 4).font = subheader_font
    return ws


def write_atm_ext_headers(ws, row: int):
    """Set column headers."""

    ws.cell(row, 1).value = "Nr."
    ws.cell(row, 2).value = "Auslöser"
    ws.cell(row, 3).value = "Bedingung"
    ws.cell(row, 4).value = "Aktion"
    ws.cell(row, 1).font = subheader_font
    ws.cell(row, 2).font = subheader_font
    ws.cell(row, 3).font = subheader_font
    ws.cell(row, 4).font = subheader_font
    return ws


def get_area_name(entity, mod) -> str:
    """Return name of area."""

    if entity.area:
        return mod.api_srv.routers[mod.rt_id - 1].get_area_name(entity.area)
    else:
        return mod.get_area_name()


def add_css_info() -> str:
    """Apply css style definition to page."""
    page = ""
    page += "<head>\n"
    page += "  <style>\n"
    page += '    body {margin-left: 60px; font: 10px/1.231 arial, helvetica, clean, sans-serif;line-height: 16px; font-family: "Lucida Sans", Arial, Verdana, Helvetica, sans-serif;}'
    page += "    h1 {padding-bottom: 5px; padding-top: 20px; font-size: 16px; color: #c0372d;}\n"
    page += "    h2 {padding-bottom: 5px; padding-top: 12px; font-size: 14px; font-weight: bold;}\n"
    page += "    hr {width: 635px; margin-left: 0px; margin-top: 10px; border-color: #c0372d;}\n"
    page += "    table {font-weight: inherit; font-style: inherit; font-size: 100%; width: 640px;}\n"
    page += "    th {background-color: #af6258; color: #ffffff; padding: 3px; padding-top: 6px; padding-bottom: 6px;text-align: left; }\n"
    page += "    tr:nth-child(odd) {background-color: #f4e5e3;}\n"
    page += "    a:-webkit-any-link {color: #c0372d; text-decoration: none;}\n"
    page += "    a:hover {text-decoration: underline;}\n"
    page += (
        "    @media print { .pagebreak { clear: both; page-break-after: always; } }\n"
    )
    page += "  </style>\n"
    page += "</head>\n"
    page += "<body>\n"
    return page


def clean_name(in_str: str) -> str:
    """Strip control characters from string."""
    return "".join(i for i in in_str if (i.isprintable() and i != "\xff"))
